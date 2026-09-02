#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Build the master Excel + CSV catalog from data/products/catalog_*.jsonl
Dedupe by SKU (first occurrence wins), keep every fetched record.
Run:  /tmp/venv/bin/python tools/build_excel.py
"""
import csv, glob, json, os, sys

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PROD_DIR = os.path.join(ROOT, "data", "products")
OUT_XLSX = os.path.join(ROOT, "data", "excel", "FS.com_catalog_US.xlsx")
OUT_CSV = os.path.join(ROOT, "data", "excel", "FS.com_catalog_US.csv")

HEADERS = [
    "SKU", "型号P/N", "一级分类", "分类路径", "商品标题", "单价USD",
    "销量原文", "销量(数值)", "评论数", "标签(热门/新品)", "库存状态",
    "可选配置", "特性标签", "主图URL", "商品页URL", "抓取日期", "来源分类页",
]

def sold_to_num(raw):
    if not raw:
        return None
    raw = raw.replace("Sold", "").strip().replace(",", "").replace(" ", "")
    if not raw:
        return None
    if raw.endswith("K") or raw.endswith("k"):
        try:
            return int(float(raw[:-1]) * 1000)
        except ValueError:
            return None
    try:
        return int(float(raw))
    except ValueError:
        return None

def load_rows():
    rows, seen = [], set()
    files = sorted(glob.glob(os.path.join(PROD_DIR, "catalog_*.jsonl")))
    if not files:
        print("no catalog_*.jsonl found in", PROD_DIR); sys.exit(1)
    for fp in files:
        with open(fp, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                r = json.loads(line)
                key = r.get("sku")
                if key in seen:
                    continue
                seen.add(key)
                r["_src"] = os.path.basename(fp)
                rows.append(r)
    return rows

def main():
    rows = load_rows()
    # order: keep file order; compute derived
    def cat_of(r):
        c = r.get("cat", "")
        return c.split(" > ")[0] if " > " in c else c
    rows.sort(key=lambda r: (cat_of(r), -(r.get("sold_num") or -1)))

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "全站商品目录"
    ws.append(HEADERS)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="C0392B")
        c.alignment = Alignment(vertical="center")
    n_cat = {}
    for r in rows:
        ws.append([
            r.get("sku"), r.get("pn"), cat_of(r), r.get("cat"), r.get("title"),
            r.get("price"), r.get("sold_raw"),
            sold_to_num(r.get("sold_raw")) if r.get("sold_num") is None else r.get("sold_num"),
            r.get("reviews"), r.get("flag", ""), r.get("stock", ""),
            r.get("options", ""), " | ".join(r.get("tags", [])),
            r.get("img"), r.get("url"), r.get("ts"), r.get("cat_url"),
        ])
        n_cat[cat_of(r)] = n_cat.get(cat_of(r), 0) + 1
    # sheet2: stats per category
    ws2 = wb.create_sheet("按分类统计")
    ws2.append(["一级分类", "商品数(SKU)"])
    for k, v in sorted(n_cat.items()):
        ws2.append([k, v])
    ws2.append(["合计", len(rows)])
    width = [10, 26, 22, 40, 60, 11, 11, 11, 8, 12, 10, 18, 60, 60, 42, 12, 52]
    for i, w in enumerate(width, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    wb.save(OUT_XLSX)

    with open(OUT_CSV, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(HEADERS)
        for r in rows:
            w.writerow([
                r.get("sku"), r.get("pn"), cat_of(r), r.get("cat"), r.get("title"),
                r.get("price"), r.get("sold_raw"),
                sold_to_num(r.get("sold_raw")) if r.get("sold_num") is None else r.get("sold_num"),
                r.get("reviews"), r.get("flag", ""), r.get("stock", ""),
                r.get("options", ""), " | ".join(r.get("tags", [])),
                r.get("img"), r.get("url"), r.get("ts"), r.get("cat_url"),
            ])
    print(f"rows={len(rows)}  xlsx={OUT_XLSX}  csv={OUT_CSV}")
    for k, v in sorted(n_cat.items()):
        print(f"  {k}: {v}")

if __name__ == "__main__":
    main()
